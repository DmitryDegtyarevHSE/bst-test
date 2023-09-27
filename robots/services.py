from datetime import datetime, timedelta
from typing import List, Dict

from django.core.mail import send_mass_mail
from django.db.models import Count
from openpyxl import Workbook

from customers.models import Customer
from orders.models import Order
from robots.models import Robot


def send_notifications(serial: str, sender_email: str, receivers: List[Customer]) -> None:
    """ Функция отправки уведомлений """
    model, version = serial.split('-')
    subject = "Робот поступил в наличие"
    message = f"Добрый день! Недавно вы интересовались нашим роботом модели {model}, версии {version}. Этот робот теперь в наличии. Если вам подходит этот вариант - пожалуйста, свяжитесь с нами"
    sender = sender_email
    datatuple = ((subject, message, sender, [receiver.email]) for receiver in receivers)
    send_mass_mail(datatuple)
    return


def get_customers_list(serial: str) -> List[Customer]:
    """ Получение списка заказчиков модели"""
    orders = Order.objects.filter(robot_serial=serial).select_related('customer')
    customers_set = set(order.customer for order in orders)
    return list(customers_set)


def _get_weekly_count() -> List:
    """ Получение роботов, произведенных за неделю """
    week_ago = datetime.now() - timedelta(days=7)
    return Robot.objects.filter(created__gte=week_ago).values('model', 'version').annotate(mcount=Count('version'))


def save_robot(data: Dict) -> None:
    """ Функция сохранения экземпляра робота """
    robot = Robot(
        serial=f"{data['model']}-{data['version']}",
        model=data['model'],
        version=data['version'],
        created=data['created']
    )
    robot.full_clean()
    robot.save()


def create_excel_report() -> Workbook:
    """ Функция создания отчета о роботах, произведенных за неделю """
    wb = Workbook()
    ws = wb.active

    ws['A1'] = 'Модель'
    ws['B1'] = 'Версия'
    ws['C1'] = 'Количество за неделю'

    robots_grouped_by_version = _get_weekly_count()
    row_num = 2
    for robot in robots_grouped_by_version:
        ws.cell(row=row_num, column=1, value=robot['model'])
        ws.cell(row=row_num, column=2, value=robot['version'])
        ws.cell(row=row_num, column=3, value=robot['mcount'])
        row_num += 1

    return wb


def get_current_date():
    """ Функция получения текущей даты """
    current_date = datetime.now()
    return f"{current_date.year}-{current_date.month}-{current_date.day}"

