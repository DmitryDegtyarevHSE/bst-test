from datetime import datetime
import json

from django.core.exceptions import ValidationError
from django.http import JsonResponse, HttpResponse
from django.utils.decorators import method_decorator
from django.views import View
from django.views.decorators.csrf import csrf_exempt
from openpyxl import Workbook

from .models import Robot


@method_decorator(csrf_exempt, name="dispatch")
class RobotView(View):
    http_method_names = ['post']

    def post(self, request):
        try:
            data = json.loads(request.body)
            robot = Robot(
                model=data["model"],
                version=data["version"],
                created=data["created"],
            )
            robot.full_clean()
            robot.save()
            return JsonResponse({"message": "created"}, status=200)
        except (KeyError, json.JSONDecodeError, ValidationError) as e:
            return JsonResponse({'error': str(e)}, status=400)


def download_excel(request):
    wb = Workbook()
    ws = wb.active

    ws['A1'] = 'Модель'
    ws['B1'] = 'Версия'
    ws['C1'] = 'Количество за неделю'

    robots_grouped_by_version = Robot.get_weekly_count()
    row_num = 2
    for robot in robots_grouped_by_version:
        ws.cell(row=row_num, column=1, value=robot['model'])
        ws.cell(row=row_num, column=2, value=robot['version'])
        ws.cell(row=row_num, column=3, value=robot['mcount'])
        row_num += 1

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = \
        f'attachment; filename=robots-count-week-{get_current_date()}.xlsx'
    wb.save(response)
    return response


def get_current_date():
    current_date = datetime.now()
    return f"{current_date.year}-{current_date.month}-{current_date.day}"
